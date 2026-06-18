import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, datetime
from pathlib import Path
from ...styles import COLORS, FONT

BG = "#f1f5f9"; CARD = "#ffffff"; BDR = "#e2e8f0"
TXT = "#1e293b"; MUTED = "#64748b"; HDR_BG = "#0f172a"
BLUE = "#3b82f6"; BLUE_H = "#2563eb"
GREEN = "#10b981"; GREEN_H = "#059669"
RED = "#ef4444"; RED_H = "#dc2626"
ORANGE = "#f59e0b"; PURPLE = "#8b5cf6"

def _btn(p, text, bg, hover, cmd=None, padx=14, pady=8):
    b = tk.Button(p, text=text, bg=bg, fg="white", font=("Segoe UI", 10, "bold"),
                  bd=0, relief="flat", padx=padx, pady=pady, cursor="hand2",
                  activebackground=hover, activeforeground="white", command=cmd)
    b.bind("<Enter>", lambda e: b.config(bg=hover))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

class StockStatementScreen(tk.Frame):
    def __init__(self, parent, app_instance, **kwargs):
        super().__init__(parent, bg=BG, **kwargs)
        self.app = app_instance
        self._rows = []
        self._var_search = tk.StringVar()
        self._var_from = tk.StringVar()
        self._var_to = tk.StringVar()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=HDR_BG, padx=20, pady=14)
        hdr.pack(fill="x")
        
        left = tk.Frame(hdr, bg=HDR_BG)
        left.pack(side="left")
        tk.Label(left, text="📋 Stock Statement Report", font=("Segoe UI", 16, "bold"),
                 fg="white", bg=HDR_BG).pack(anchor="w")
        tk.Label(left, text="Opening • Received • Sold • Balance", font=("Segoe UI", 9),
                 fg="#94a3b8", bg=HDR_BG).pack(anchor="w")
        
        right = tk.Frame(hdr, bg=HDR_BG)
        right.pack(side="right")
        
        # View Mode Toggle
        mode_frame = tk.Frame(right, bg="#1e3a5f", relief="flat")
        mode_frame.pack(side="left", padx=(0, 10))
        
        tk.Label(mode_frame, text="View:", font=("Segoe UI", 9, "bold"),
                 fg="#93c5fd", bg="#1e3a5f", padx=8, pady=6).pack(side="left")
        
        self._mode_var = tk.StringVar(value="current")
        
        tk.Radiobutton(
            mode_frame, text="Current Stock", variable=self._mode_var, value="current",
            font=("Segoe UI", 9), bg="#1e3a5f", fg="white", selectcolor="#1e3a5f",
            activebackground="#2563eb", activeforeground="white",
            command=self.on_show, cursor="hand2"
        ).pack(side="left", padx=4)
        
        tk.Radiobutton(
            mode_frame, text="FY Movement", variable=self._mode_var, value="fy",
            font=("Segoe UI", 9), bg="#1e3a5f", fg="white", selectcolor="#1e3a5f",
            activebackground="#2563eb", activeforeground="white",
            command=self.on_show, cursor="hand2"
        ).pack(side="left", padx=4)
        
        _btn(right, "📊 Excel", GREEN, GREEN_H, self._export_excel).pack(side="left", padx=3)

        tk.Frame(self, bg=BDR, height=1).pack(fill="x")

        # Stats
        self._stat_vars = {}
        self._fy_label = tk.StringVar(value="FY 2024-25")
        stat_row = tk.Frame(self, bg=BG, padx=16, pady=14)
        stat_row.pack(fill="x")
        
        stats = [("Opening", "opening", PURPLE), ("Received", "received", GREEN),
                 ("Sold", "sold", RED), ("Balance", "balance", BLUE),
                 ("Value", "value", ORANGE)]
        
        for idx, (label, key, color) in enumerate(stats):
            c = tk.Frame(stat_row, bg="#fafafa", highlightbackground=BDR, highlightthickness=1)
            c.pack(side="left", expand=True, fill="x", padx=5)
            top = tk.Frame(c, bg="#fafafa", padx=14, pady=10)
            top.pack(fill="x")
            tk.Label(top, text=label, font=("Segoe UI", 8, "bold"), fg=MUTED, bg="#fafafa").pack(anchor="w")
            var = tk.StringVar(value="0")
            self._stat_vars[key] = var
            tk.Label(top, textvariable=var, font=("Segoe UI", 16, "bold"), fg=TXT, bg="#fafafa").pack(anchor="w")
            tk.Frame(c, bg=color, height=3).pack(fill="x", side="bottom")

        # Filters
        flt = tk.Frame(self, bg=CARD, padx=20, pady=14, highlightbackground=BDR, highlightthickness=1)
        flt.pack(fill="x", padx=16)
        
        # Create inner frame for proper layout
        flt_inner = tk.Frame(flt, bg=CARD)
        flt_inner.pack(fill="x")
        
        # Search field
        search_col = tk.Frame(flt_inner, bg=CARD)
        search_col.pack(side="left", padx=(0, 16))
        tk.Label(search_col, text="Search", font=("Segoe UI", 9), fg=MUTED, bg=CARD).pack(anchor="w")
        tk.Entry(search_col, textvariable=self._var_search, font=("Segoe UI", 10), width=25,
                 relief="solid", bd=1).pack(pady=3)
        
        # From Date field
        from_col = tk.Frame(flt_inner, bg=CARD)
        from_col.pack(side="left", padx=(0, 16))
        tk.Label(from_col, text="From Date", font=("Segoe UI", 9), fg=MUTED, bg=CARD).pack(anchor="w")
        from ..purchase.purchase_invoice_dialog import DateEntry as DE
        self._e_from = DE(from_col, textvariable=self._var_from, width=14, bg=CARD)
        self._e_from.pack(pady=3)
        
        # To Date field
        to_col = tk.Frame(flt_inner, bg=CARD)
        to_col.pack(side="left", padx=(0, 16))
        tk.Label(to_col, text="To Date", font=("Segoe UI", 9), fg=MUTED, bg=CARD).pack(anchor="w")
        self._e_to = DE(to_col, textvariable=self._var_to, width=14, bg=CARD)
        self._e_to.pack(pady=3)
        
        # Button
        btn_col = tk.Frame(flt_inner, bg=CARD)
        btn_col.pack(side="left", pady=(18, 0))
        _btn(btn_col, "🔍 Show", BLUE, BLUE_H, self.on_show, pady=7).pack()

        # Table
        tbl = tk.Frame(self, bg=CARD, highlightbackground=BDR, highlightthickness=1)
        tbl.pack(fill="both", expand=True, padx=16, pady=(10,16))
        
        tbar = tk.Frame(tbl, bg=CARD, padx=16, pady=10)
        tbar.pack(fill="x")
        tk.Frame(tbar, bg=BDR, height=1).pack(fill="x", side="bottom")
        
        tbar_left = tk.Frame(tbar, bg=CARD)
        tbar_left.pack(side="left")
        tk.Label(tbar_left, text="Stock Statement",
                 font=("Segoe UI", 11, "bold"), fg=TXT, bg=CARD).pack(side="left")
        self._fy_badge = tk.Label(tbar_left, textvariable=self._fy_label,
                                  font=("Segoe UI", 8, "bold"), fg="white",
                                  bg=BLUE, padx=6, pady=2)
        self._fy_badge.pack(side="left", padx=8)
        
        cols = ("Product", "Company", "Opening", "Received", "Sold", "Balance", "Value")
        widths = (250, 150, 120, 120, 120, 120, 130)
        
        style = ttk.Style()
        style.configure("SS.Treeview", font=("Segoe UI", 10), rowheight=36,
                        background=CARD, fieldbackground=CARD, foreground=TXT, borderwidth=0)
        style.configure("SS.Treeview.Heading", font=("Segoe UI", 9, "bold"),
                        background="#f8fafc", foreground=MUTED, relief="flat")
        
        tree_frame = tk.Frame(tbl, bg=CARD)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                  yscrollcommand=vsb.set, selectmode="browse", style="SS.Treeview")
        vsb.config(command=self.tree.yview)
        
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col.upper(), anchor="w")
            self.tree.column(col, width=w, anchor="w" if col in ("Product", "Company") else "center")
        
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

    def on_show(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        
        term = self._var_search.get().strip()
        frm = self._var_from.get().strip()
        to = self._var_to.get().strip()
        
        # Get FY dates and mode
        fy_start = getattr(self.app, 'current_fy_start', None)
        fy_end = getattr(self.app, 'current_fy_end', None)
        mode = getattr(self, '_mode_var', None)
        current_mode = mode.get() if mode else "current"
        
        # Update FY badge
        if hasattr(self.app, 'fy_var') and self.app.fy_var:
            fy_text = self.app.fy_var.get()
            self._fy_label.set(f"FY {fy_text}")
        
        # If FY mode and no custom dates, use FY dates
        if current_mode == "fy" and not frm and not to and fy_start and fy_end:
            frm = fy_start
            to = fy_end
        
        rid = self.app.config_data.retailer_id
        
        try:
            self._rows = self._fetch_stock(term, frm, to, rid)
            
            t_open = t_recv = t_sold = t_bal = t_val = 0.0
            t_open_f = t_recv_f = t_sold_f = t_bal_f = 0.0
            
            for idx, r in enumerate(self._rows):
                o = r['opening']; of = r['opening_free']
                rc = r['received']; rf = r['received_free']
                s = r['sold']; sf = r['sold_free']
                b = r['balance']; bf = r['balance_free']
                v = r['value']
                
                t_open += o; t_open_f += of
                t_recv += rc; t_recv_f += rf
                t_sold += s; t_sold_f += sf
                t_bal += b; t_bal_f += bf
                t_val += v
                
                tag = "odd" if idx % 2 else "even"
                self.tree.insert("", "end", tags=(tag,), values=(
                    r['product_name'], r['company'],
                    f"{o:.0f}+{of:.0f}F" if of else f"{o:.0f}",
                    f"{rc:.0f}+{rf:.0f}F" if rf else f"{rc:.0f}",
                    f"{s:.0f}+{sf:.0f}F" if sf else f"{s:.0f}",
                    f"{b:.0f}+{bf:.0f}F" if bf else f"{b:.0f}",
                    f"₹{v:,.2f}"
                ))
            
            self._stat_vars["opening"].set(f"{t_open:.0f}+{t_open_f:.0f}F")
            self._stat_vars["received"].set(f"{t_recv:.0f}+{t_recv_f:.0f}F")
            self._stat_vars["sold"].set(f"{t_sold:.0f}+{t_sold_f:.0f}F")
            self._stat_vars["balance"].set(f"{t_bal:.0f}+{t_bal_f:.0f}F")
            self._stat_vars["value"].set(f"₹{t_val:,.2f}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {e}", parent=self)

    def _fetch_stock(self, term, frm, to, rid):
        import re
        
        # Helper to extract packing multiplier
        def get_packing_mult(packing_str):
            if not packing_str: return 1
            m = re.search(r'(\d+)[xX*](\d+)', str(packing_str))
            if m: return int(m.group(1)) * int(m.group(2))
            m2 = re.search(r'(\d+)', str(packing_str))
            return int(m2.group(1)) if m2 else 1
        
        # Build queries with date filters
        w_pur = "WHERE pm.retailer_id=%s"
        p_pur = [rid]
        if term:
            w_pur += " AND (p.product_name LIKE %s OR p.product_company LIKE %s)"
            p_pur += [f"%{term}%", f"%{term}%"]
        
        # Opening stock: purchases BEFORE from date
        opening_data = {}
        if frm:
            w_open_pur = w_pur + " AND i.invoice_date < %s"
            p_open_pur = p_pur + [frm]
            
            open_purchases = self.app.db.query(f"""
                SELECT p.productid, p.product_packing,
                       SUM(pm.product_quantity) AS qty,
                       SUM(pm.product_free_qty) AS free_qty
                FROM core_purchasemaster pm
                JOIN core_productmaster p ON p.productid = pm.productid_id
                JOIN core_invoicemaster i ON i.invoiceid = pm.product_invoiceid_id
                {w_open_pur}
                GROUP BY p.productid, p.product_packing
            """, p_open_pur)
            
            w_open_sal = "WHERE sm.retailer_id=%s"
            p_open_sal = [rid]
            if term:
                w_open_sal += " AND (p.product_name LIKE %s OR p.product_company LIKE %s)"
                p_open_sal += [f"%{term}%", f"%{term}%"]
            w_open_sal += " AND si.sales_invoice_date < %s"
            p_open_sal.append(frm)
            
            open_sales = self.app.db.query(f"""
                SELECT sm.productid_id AS productid,
                       SUM(sm.sale_quantity) AS qty,
                       SUM(sm.sale_free_qty) AS free_qty
                FROM core_salesmaster sm
                JOIN core_salesinvoicemaster si ON si.sales_invoice_no = sm.sales_invoice_no_id
                JOIN core_productmaster p ON p.productid = sm.productid_id
                {w_open_sal}
                GROUP BY sm.productid_id
            """, p_open_sal)
            
            open_sales_dict = {s['productid']: {'qty': float(s.get('qty', 0) or 0),
                                                'free_qty': float(s.get('free_qty', 0) or 0)} for s in open_sales}
            
            for op in open_purchases:
                pid = op['productid']
                mult = get_packing_mult(op.get('product_packing', ''))
                op_qty = float(op.get('qty', 0) or 0) * mult
                op_free = float(op.get('free_qty', 0) or 0)
                s = open_sales_dict.get(pid, {'qty': 0, 'free_qty': 0})
                opening_data[pid] = {
                    'qty': op_qty - s['qty'],
                    'free_qty': op_free - s['free_qty']
                }
        
        # Period purchases
        if frm:
            w_pur += " AND i.invoice_date >= %s"
            p_pur.append(frm)
        if to:
            w_pur += " AND i.invoice_date <= %s"
            p_pur.append(to)
        
        purchases = self.app.db.query(f"""
            SELECT p.productid, p.product_name, p.product_company, p.product_packing,
                   SUM(pm.product_quantity) AS qty,
                   SUM(pm.product_free_qty) AS free_qty,
                   MAX(pm.product_MRP) AS mrp
            FROM core_purchasemaster pm
            JOIN core_productmaster p ON p.productid = pm.productid_id
            JOIN core_invoicemaster i ON i.invoiceid = pm.product_invoiceid_id
            {w_pur}
            GROUP BY p.productid, p.product_name, p.product_company, p.product_packing
            ORDER BY p.product_name
        """, p_pur)
        
        # Period sales
        w_sal = "WHERE sm.retailer_id=%s"
        p_sal = [rid]
        if term:
            w_sal += " AND (p.product_name LIKE %s OR p.product_company LIKE %s)"
            p_sal += [f"%{term}%", f"%{term}%"]
        if frm:
            w_sal += " AND si.sales_invoice_date >= %s"
            p_sal.append(frm)
        if to:
            w_sal += " AND si.sales_invoice_date <= %s"
            p_sal.append(to)
        
        sales = self.app.db.query(f"""
            SELECT sm.productid_id AS productid,
                   SUM(sm.sale_quantity) AS qty,
                   SUM(sm.sale_free_qty) AS free_qty
            FROM core_salesmaster sm
            JOIN core_salesinvoicemaster si ON si.sales_invoice_no = sm.sales_invoice_no_id
            JOIN core_productmaster p ON p.productid = sm.productid_id
            {w_sal}
            GROUP BY sm.productid_id
        """, p_sal)
        
        sales_dict = {s['productid']: {'qty': float(s.get('qty', 0) or 0),
                                        'free_qty': float(s.get('free_qty', 0) or 0)} for s in sales}
        
        # Stock issues (reductions) - safely check if table exists
        issues_dict = {}
        try:
            # First check if table exists
            table_check = self.app.db.query("""
                SELECT COUNT(*) as cnt FROM information_schema.tables 
                WHERE table_schema = DATABASE() 
                AND table_name = 'core_stockissuedetail'
            """, ())
            
            if table_check and table_check[0].get('cnt', 0) > 0:
                w_issue = "WHERE sid.retailer_id=%s"
                p_issue = [rid]
                if term:
                    w_issue += " AND (p.product_name LIKE %s OR p.product_company LIKE %s)"
                    p_issue += [f"%{term}%", f"%{term}%"]
                if frm:
                    w_issue += " AND si.issue_date >= %s"
                    p_issue.append(frm)
                if to:
                    w_issue += " AND si.issue_date <= %s"
                    p_issue.append(to)
                
                issues = self.app.db.query(f"""
                    SELECT sid.product_id AS productid,
                           SUM(sid.quantity_issued) AS qty
                    FROM core_stockissuedetail sid
                    JOIN core_stockissuemaster si ON si.issue_id = sid.issue_id
                    JOIN core_productmaster p ON p.productid = sid.product_id
                    {w_issue}
                    GROUP BY sid.product_id
                """, p_issue)
                
                issues_dict = {i['productid']: float(i.get('qty', 0) or 0) for i in issues}
        except Exception as e:
            # Silently skip if table doesn't exist
            import logging
            logging.debug(f"Stock issues table not available: {e}")
        
        result = []
        for p in purchases:
            pid = p['productid']
            mult = get_packing_mult(p.get('product_packing', ''))
            
            # Opening
            opening = opening_data.get(pid, {'qty': 0, 'free_qty': 0})
            open_q = opening['qty']
            open_f = opening['free_qty']
            
            # Received (qty * inpacking, free qty as-is)
            recv_q = float(p.get('qty', 0) or 0) * mult
            recv_f = float(p.get('free_qty', 0) or 0)
            
            # Sold
            s = sales_dict.get(pid, {'qty': 0, 'free_qty': 0})
            sold_q = s['qty']
            sold_f = s['free_qty']
            
            # Issues
            issued_q = issues_dict.get(pid, 0)
            
            # Balance = Opening + Received - Sold - Issued
            bal_q = open_q + recv_q - sold_q - issued_q
            bal_f = open_f + recv_f - sold_f
            
            mrp = float(p.get('mrp', 0) or 0)
            val = bal_q * mrp
            
            result.append({
                'product_name': p['product_name'],
                'company': p.get('product_company', ''),
                'opening': open_q,
                'opening_free': open_f,
                'received': recv_q,
                'received_free': recv_f,
                'sold': sold_q,
                'sold_free': sold_f,
                'balance': bal_q,
                'balance_free': bal_f,
                'value': val
            })
        
        return result

    def _export_excel(self):
        if not self._rows:
            messagebox.showinfo("No Data", "Load data first.", parent=self)
            return
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             initialfile=f"StockStatement_{date.today()}.xlsx",
                                             parent=self)
        if not path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Statement"
            
            ws['A1'] = "STOCK STATEMENT REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:G1')
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(size=9)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:G2')
            
            hdrs = ["Product", "Company", "Opening", "Received", "Sold", "Balance", "Value"]
            hdr_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF")
            
            for col, h in enumerate(hdrs, 1):
                c = ws.cell(4, col, h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal='center')
            
            row = 5
            for r in self._rows:
                ws.cell(row, 1, r['product_name'])
                ws.cell(row, 2, r['company'])
                ws.cell(row, 3, f"{r['opening']:.0f}+{r['opening_free']:.0f}F" if r['opening_free'] else f"{r['opening']:.0f}")
                ws.cell(row, 4, f"{r['received']:.0f}+{r['received_free']:.0f}F" if r['received_free'] else f"{r['received']:.0f}")
                ws.cell(row, 5, f"{r['sold']:.0f}+{r['sold_free']:.0f}F" if r['sold_free'] else f"{r['sold']:.0f}")
                ws.cell(row, 6, f"{r['balance']:.0f}+{r['balance_free']:.0f}F" if r['balance_free'] else f"{r['balance']:.0f}")
                ws.cell(row, 7, r['value'])
                row += 1
            
            wb.save(path)
            messagebox.showinfo("Exported", f"Saved:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}", parent=self)

    def kb_refresh(self):
        self.on_show()
    
    def kb_search(self):
        self._var_search.focus_set()
