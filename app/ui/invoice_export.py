"""invoice_export.py — Professional PDF & Excel export for Purchase and Sales invoices."""
from __future__ import annotations

import os
from pathlib import Path
from typing import Any

# ── openpyxl ──────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

# ── reportlab ─────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT


# ── helpers ───────────────────────────────────────────────────────────────────

def _out_dir() -> Path:
    # Use Documents folder to avoid Windows Defender blocking exports from program directories
    docs = Path.home() / "Documents" / "MedicVista_Exports"
    docs.mkdir(parents=True, exist_ok=True)
    return docs


def _s(v: Any) -> str:
    return "" if v is None else str(v)


def _m(v: Any) -> str:
    try:
        return f"₹{float(v):.2f}"
    except Exception:
        return _s(v)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _safe_float(v: Any, fallback: float = 0.0) -> float:
    if v is None:
        return fallback
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace("₹", "").replace(",", "").strip()
        if not s:
            return fallback
        return float(s)
    except ValueError:
        return fallback


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORTS
# ═══════════════════════════════════════════════════════════════════════════════

def _excel_header(ws, title: str, meta: list[tuple[str, str]]):
    """Write a styled header block."""
    ws.merge_cells("A1:N1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1e293b")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for i, (k, v) in enumerate(meta, 3):
        ws.cell(i, 1, k).font = Font(bold=True, color="475569")
        ws.cell(i, 2, v)
    return 3 + len(meta) + 1  # return next row


def _excel_item_headers(ws, headers: list[str], row: int,
                         fill_color="1e3a5f"):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 22


def _excel_data_row(ws, row: int, values: list, alt: bool = False):
    fill = PatternFill("solid", fgColor="F1F5F9") if alt else None
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill


def _excel_totals_row(ws, row: int, values: list):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="right" if isinstance(v, float) else "left",
                                    vertical="center")


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                best = max(best, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)


# ── Sales Excel ───────────────────────────────────────────────────────────────

def export_sales_invoice_excel(db, invoice_no: str, out_dir: Path | None = None) -> str:
    out_dir = out_dir or _out_dir()
    inv   = db.get_sales_invoice(invoice_no)
    if not inv:
        raise ValueError(f"Sales invoice not found: {invoice_no}")
    items = db.get_sales_items(invoice_no)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Invoice"

    meta = [
        ("Invoice No",  _s(invoice_no)),
        ("Date",        _s(inv.get("sales_invoice_date"))),
        ("Customer",    _s(inv.get("customer_name"))),
        ("Type",        _s(inv.get("customer_type", ""))),
        ("Transport",   _m(inv.get("sales_transport_charges", 0))),
    ]
    data_row = _excel_header(ws, f"MedicVista — Sales Invoice: {invoice_no}", meta)

    headers = ["#", "Product", "Company", "Packing", "Batch", "Expiry",
               "MRP", "Rate", "Qty", "Free", "Disc", "CGST%", "SGST%", "Total"]
    _excel_item_headers(ws, headers, data_row)
    data_row += 1

    subtotal = 0.0
    for i, it in enumerate(items, 1):
        total = float(it.get("sale_total_amount") or 0)
        subtotal += total
        _excel_data_row(ws, data_row, [
            i,
            _s(it.get("product_name") or it.get("prod_name")),
            _s(it.get("product_company")),
            _s(it.get("product_packing")),
            _s(it.get("product_batch_no")),
            _s(it.get("product_expiry")),
            float(it.get("product_MRP") or 0),
            float(it.get("sale_rate") or 0),
            float(it.get("sale_quantity") or 0),
            float(it.get("sale_free_qty") or 0),
            float(it.get("sale_discount") or 0),
            float(it.get("sale_cgst") or 0),
            float(it.get("sale_sgst") or 0),
            total,
        ], alt=(i % 2 == 0))
        data_row += 1

    transport = float(inv.get("sales_transport_charges") or 0)
    grand     = subtotal + transport
    _excel_totals_row(ws, data_row, ["", "SUBTOTAL", "", "", "", "", "", "", "", "", "", "", "", subtotal])
    _excel_totals_row(ws, data_row + 1, ["", "TRANSPORT", "", "", "", "", "", "", "", "", "", "", "", transport])
    _excel_totals_row(ws, data_row + 2, ["", "GRAND TOTAL", "", "", "", "", "", "", "", "", "", "", "", grand])

    _auto_width(ws)
    path = out_dir / f"Sales_{invoice_no}.xlsx"
    wb.save(str(path))
    return str(path)


# ── Purchase Excel ────────────────────────────────────────────────────────────

def export_purchase_invoice_excel(db, invoice_no: str, out_dir: Path | None = None) -> str:
    out_dir = out_dir or _out_dir()
    rows = db.query(
        """SELECT i.invoiceid, i.invoice_no, i.invoice_date, s.supplier_name,
                  i.invoice_total, i.invoice_paid, i.transport_charges, i.payment_status
           FROM core_invoicemaster i
           JOIN core_suppliermaster s ON s.supplierid = i.supplierid_id
           WHERE i.invoice_no=%s""", (invoice_no,))
    if not rows:
        raise ValueError(f"Purchase invoice not found: {invoice_no}")
    inv   = rows[0]
    items = db.get_invoice_items(int(inv["invoiceid"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Invoice"

    meta = [
        ("Invoice No",  _s(invoice_no)),
        ("Date",        _s(inv.get("invoice_date"))),
        ("Supplier",    _s(inv.get("supplier_name"))),
        ("Transport",   _m(inv.get("transport_charges", 0))),
        ("Status",      _s(inv.get("payment_status", "")).upper()),
    ]
    data_row = _excel_header(ws, f"MedicVista — Purchase Invoice: {invoice_no}", meta)

    headers = ["#", "Product", "Company", "Packing", "Batch", "Expiry",
               "MRP", "Pur.Rate", "Qty", "Free", "Disc", "CGST%", "SGST%", "Total"]
    _excel_item_headers(ws, headers, data_row, fill_color="7c3aed")
    data_row += 1

    subtotal = 0.0
    for i, it in enumerate(items, 1):
        total = float(it.get("total_amount") or 0)
        subtotal += total
        _excel_data_row(ws, data_row, [
            i,
            _s(it.get("prod_name") or it.get("product_name")),
            _s(it.get("product_company")),
            _s(it.get("product_packing")),
            _s(it.get("product_batch_no")),
            _s(it.get("product_expiry")),
            float(it.get("product_MRP") or 0),
            float(it.get("product_purchase_rate") or 0),
            float(it.get("product_quantity") or 0),
            float(it.get("product_free_qty") or 0),
            float(it.get("product_discount_got") or 0),
            float(it.get("CGST") or 0),
            float(it.get("SGST") or 0),
            total,
        ], alt=(i % 2 == 0))
        data_row += 1

    transport = float(inv.get("transport_charges") or 0)
    grand     = subtotal + transport
    _excel_totals_row(ws, data_row,     ["", "SUBTOTAL", "", "", "", "", "", "", "", "", "", "", "", subtotal])
    _excel_totals_row(ws, data_row + 1, ["", "TRANSPORT", "", "", "", "", "", "", "", "", "", "", "", transport])
    _excel_totals_row(ws, data_row + 2, ["", "GRAND TOTAL", "", "", "", "", "", "", "", "", "", "", "", grand])

    _auto_width(ws)
    path = out_dir / f"Purchase_{invoice_no}.xlsx"
    wb.save(str(path))
    return str(path)


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORTS
# ═══════════════════════════════════════════════════════════════════════════════

_STYLES = getSampleStyleSheet()

def _pdf_style(name, parent="Normal", **kw):
    return ParagraphStyle(name, parent=_STYLES[parent], **kw)


_H1    = _pdf_style("h1",   fontSize=16, textColor=colors.white,
                             leading=20, alignment=TA_CENTER)
_H2    = _pdf_style("h2",   fontSize=11, textColor=colors.HexColor("#1e293b"),
                             leading=14, fontName="Helvetica-Bold")
_BODY  = _pdf_style("body", fontSize=9,  textColor=colors.HexColor("#334155"),
                             leading=12)
_SMALL = _pdf_style("sm",   fontSize=8,  textColor=colors.HexColor("#64748b"),
                             leading=10)
_BOLD  = _pdf_style("bold", fontSize=9,  textColor=colors.HexColor("#1e293b"),
                             fontName="Helvetica-Bold", leading=12)
_RIGHT = _pdf_style("right", fontSize=9, alignment=TA_RIGHT,
                              textColor=colors.HexColor("#1e293b"), leading=12)

# table style helpers
def _tbl_style(header_color=colors.HexColor("#1e293b")):
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F1F5F9")]),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ])


def _totals_block(data: list[tuple[str, str]], width):
    """Right-aligned 2-col totals mini-table."""
    tbl = Table(data, colWidths=[width * 0.7, width * 0.3])
    tbl.setStyle(TableStyle([
        ("ALIGN",       (0, 0), (0, -1), "RIGHT"),
        ("ALIGN",       (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#DBEAFE")),
        ("LINEABOVE",   (0, -1), (-1, -1), 0.8, colors.HexColor("#3B82F6")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
    ]))
    return tbl


# ── Sales PDF ─────────────────────────────────────────────────────────────────

def export_sales_invoice_pdf(db, invoice_no: str, out_dir: Path | None = None,
                              output_path=None) -> str:
    out_dir = out_dir or _out_dir()
    path = Path(output_path) if output_path else out_dir / f"Sales_{invoice_no}.pdf"

    inv   = db.get_sales_invoice(invoice_no)
    if not inv:
        raise ValueError(f"Sales invoice not found: {invoice_no}")
    items = db.get_sales_items(invoice_no)

    W, H  = A4
    doc   = SimpleDocTemplate(str(path), pagesize=A4,
                               leftMargin=15*mm, rightMargin=15*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    usable = W - 30*mm
    story  = []

    # ── Title banner ──────────────────────────────────────────────────────
    banner = Table([[Paragraph(f"MedicVista — Sales Invoice", _H1)]], colWidths=[usable])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1e293b")),
        ("ROUNDEDCORNERS", [4]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
    ]))
    story.append(banner)
    story.append(Spacer(1, 6*mm))

    # ── Info grid ─────────────────────────────────────────────────────────
    left_info = [
        [Paragraph("<b>Invoice No:</b>", _BODY),
         Paragraph(_s(invoice_no), _BOLD)],
        [Paragraph("<b>Date:</b>", _BODY),
         Paragraph(_s(inv.get("sales_invoice_date")), _BOLD)],
        [Paragraph("<b>Customer:</b>", _BODY),
         Paragraph(_s(inv.get("customer_name")), _BOLD)],
        [Paragraph("<b>Type:</b>", _BODY),
         Paragraph(_s(inv.get("customer_type", "")), _BODY)],
    ]
    info_tbl = Table(left_info, colWidths=[usable * 0.2, usable * 0.8])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX",        (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Items table ───────────────────────────────────────────────────────
    col_w = [usable * x for x in [0.22, 0.08, 0.08, 0.08, 0.07,
                                    0.07, 0.07, 0.06, 0.06, 0.06, 0.09, 0.06]]
    hdr = ["Product", "Batch", "Expiry", "MRP", "Rate",
           "Qty", "Free", "Disc", "CGST%", "SGST%", "Total", "#"]
    tdata = [hdr]
    subtotal = 0.0
    for i, it in enumerate(items, 1):
        total = float(it.get("sale_total_amount") or 0)
        subtotal += total
        tdata.append([
            _s(it.get("product_name") or it.get("prod_name")),
            _s(it.get("product_batch_no")),
            _s(it.get("product_expiry")),
            _m(it.get("product_MRP")),
            _m(it.get("sale_rate")),
            _s(it.get("sale_quantity")),
            _s(it.get("sale_free_qty")),
            _m(it.get("sale_discount")),
            _s(it.get("sale_cgst")),
            _s(it.get("sale_sgst")),
            _m(total),
            str(i),
        ])

    tbl = Table(tdata, colWidths=col_w, repeatRows=1)
    tbl.setStyle(_tbl_style(colors.HexColor("#0f172a")))
    story.append(tbl)
    story.append(Spacer(1, 4*mm))

    # ── Totals ────────────────────────────────────────────────────────────
    transport = float(inv.get("sales_transport_charges") or 0)
    grand     = subtotal + transport
    paid      = float(inv.get("sales_invoice_paid") or 0)
    balance   = grand - paid
    totals_data = [
        [Paragraph("Subtotal:", _RIGHT), Paragraph(_m(subtotal), _BOLD)],
        [Paragraph("Transport:", _RIGHT), Paragraph(_m(transport), _BOLD)],
        [Paragraph("Grand Total:", _RIGHT), Paragraph(_m(grand), _BOLD)],
        [Paragraph("Paid:", _RIGHT), Paragraph(_m(paid), _BOLD)],
        [Paragraph("Balance Due:", _RIGHT),
         Paragraph(f'<font color="red">{_m(balance)}</font>' if balance > 0.01
                   else _m(balance), _BOLD)],
    ]
    story.append(_totals_block(totals_data, usable))

    # ── Footer ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width=usable, thickness=0.5,
                              color=colors.HexColor("#E2E8F0")))
    story.append(Paragraph("Thank you for your business — MedicVista Pharmacy ERP",
                            _SMALL))

    doc.build(story)
    return str(path)


# ── Purchase PDF ──────────────────────────────────────────────────────────────

def export_purchase_invoice_pdf(db, invoice_no: str, out_dir: Path | None = None,
                                 output_path=None) -> str:
    out_dir = out_dir or _out_dir()
    path = Path(output_path) if output_path else out_dir / f"Purchase_{invoice_no}.pdf"

    rows = db.query(
        """SELECT i.invoiceid, i.invoice_no, i.invoice_date, s.supplier_name,
                  s.supplier_mobile, s.supplier_gstno,
                  i.invoice_total, i.invoice_paid, i.transport_charges, i.payment_status
           FROM core_invoicemaster i
           JOIN core_suppliermaster s ON s.supplierid = i.supplierid_id
           WHERE i.invoice_no=%s""", (invoice_no,))
    if not rows:
        raise ValueError(f"Purchase invoice not found: {invoice_no}")
    inv   = rows[0]
    items = db.get_invoice_items(int(inv["invoiceid"]))

    W, H  = A4
    doc   = SimpleDocTemplate(str(path), pagesize=A4,
                               leftMargin=15*mm, rightMargin=15*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    usable = W - 30*mm
    story  = []

    # Title
    banner = Table([[Paragraph("MedicVista — Purchase Invoice", _H1)]], colWidths=[usable])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#7c3aed")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
    ]))
    story.append(banner)
    story.append(Spacer(1, 6*mm))

    # Info
    left_info = [
        [Paragraph("<b>Invoice No:</b>", _BODY),
         Paragraph(_s(invoice_no), _BOLD)],
        [Paragraph("<b>Date:</b>", _BODY),
         Paragraph(_s(inv.get("invoice_date")), _BOLD)],
        [Paragraph("<b>Supplier:</b>", _BODY),
         Paragraph(_s(inv.get("supplier_name")), _BOLD)],
        [Paragraph("<b>GST:</b>", _BODY),
         Paragraph(_s(inv.get("supplier_gstno", "")), _BODY)],
        [Paragraph("<b>Status:</b>", _BODY),
         Paragraph(_s(inv.get("payment_status", "")).upper(), _BOLD)],
    ]
    info_tbl = Table(left_info, colWidths=[usable * 0.2, usable * 0.8])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F3FF")),
        ("BOX",        (0, 0), (-1, -1), 0.5, colors.HexColor("#DDD6FE")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 5*mm))

    # Items
    col_w = [usable * x for x in [0.20, 0.08, 0.08, 0.07, 0.07,
                                    0.07, 0.06, 0.06, 0.06, 0.06, 0.06, 0.09, 0.08]]
    hdr = ["Product", "Batch", "Expiry", "MRP", "Pur.Rate",
           "Qty", "Free", "Disc", "CGST%", "SGST%", "RateA", "Total", "#"]
    tdata = [hdr]
    subtotal = 0.0
    for i, it in enumerate(items, 1):
        total = float(it.get("total_amount") or 0)
        subtotal += total
        tdata.append([
            _s(it.get("prod_name") or it.get("product_name")),
            _s(it.get("product_batch_no")),
            _s(it.get("product_expiry")),
            _m(it.get("product_MRP")),
            _m(it.get("product_purchase_rate")),
            _s(it.get("product_quantity")),
            _s(it.get("product_free_qty")),
            _m(it.get("product_discount_got")),
            _s(it.get("CGST")),
            _s(it.get("SGST")),
            _m(it.get("rate_a")),
            _m(total),
            str(i),
        ])

    tbl = Table(tdata, colWidths=col_w, repeatRows=1)
    tbl.setStyle(_tbl_style(colors.HexColor("#7c3aed")))
    story.append(tbl)
    story.append(Spacer(1, 4*mm))

    transport = float(inv.get("transport_charges") or 0)
    grand     = subtotal + transport
    paid      = float(inv.get("invoice_paid") or 0)
    balance   = grand - paid
    totals_data = [
        [Paragraph("Subtotal:", _RIGHT), Paragraph(_m(subtotal), _BOLD)],
        [Paragraph("Transport:", _RIGHT), Paragraph(_m(transport), _BOLD)],
        [Paragraph("Grand Total:", _RIGHT), Paragraph(_m(grand), _BOLD)],
        [Paragraph("Paid:", _RIGHT), Paragraph(_m(paid), _BOLD)],
        [Paragraph("Balance Due:", _RIGHT),
         Paragraph(f'<font color="red">{_m(balance)}</font>' if balance > 0.01
                   else _m(balance), _BOLD)],
    ]
    story.append(_totals_block(totals_data, usable))

    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width=usable, thickness=0.5,
                              color=colors.HexColor("#E2E8F0")))
    story.append(Paragraph("Generated by MedicVista Pharmacy ERP", _SMALL))

    doc.build(story)
    return str(path)


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORT EXPORTS (for purchase/sales report screens)
# ═══════════════════════════════════════════════════════════════════════════════

def export_report_excel(title: str, headers: list[str],
                         rows: list[list], totals_row: list,
                         path: Path) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    data_row = _excel_header(ws, f"MedicVista — {title}", [])

    _excel_item_headers(ws, headers, data_row)
    data_row += 1
    for i, row in enumerate(rows):
        _excel_data_row(ws, data_row, row, alt=(i % 2 == 0))
        data_row += 1
    _excel_totals_row(ws, data_row, totals_row)
    _auto_width(ws)
    wb.save(str(path))
    return str(path)


def export_report_pdf(title: str, headers: list[str],
                       rows: list[list], totals_row: list,
                       summary: dict[str, str],
                       path: Path) -> str:
    W, H  = A4
    doc   = SimpleDocTemplate(str(path), pagesize=A4,
                               leftMargin=15*mm, rightMargin=15*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    usable = W - 30*mm
    story  = []

    # header
    banner = Table([[Paragraph(f"MedicVista — {title}", _H1)]], colWidths=[usable])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1e293b")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
    ]))
    story.append(banner)
    story.append(Spacer(1, 5*mm))

    # summary cards (inline)
    if summary:
        sdata = [[Paragraph(f"<b>{k}</b>: {v}", _BODY) for k, v in summary.items()]]
        stbl = Table(sdata, colWidths=[usable / len(summary)] * len(summary))
        stbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
            ("BOX",        (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",  (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(stbl)
        story.append(Spacer(1, 4*mm))

    # data table
    n = len(headers)
    col_w = [usable / n] * n
    tdata = [headers] + rows + [totals_row]
    tbl = Table(tdata, colWidths=col_w, repeatRows=1)
    ts = _tbl_style()
    ts.add("FONTNAME",    (0, len(tdata) - 1), (-1, -1), "Helvetica-Bold")
    ts.add("BACKGROUND",  (0, len(tdata) - 1), (-1, -1), colors.HexColor("#DBEAFE"))
    tbl.setStyle(ts)
    story.append(tbl)

    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=usable, thickness=0.5,
                              color=colors.HexColor("#E2E8F0")))
    story.append(Paragraph("Generated by MedicVista Pharmacy ERP", _SMALL))
    doc.build(story)
    return str(path)


# ── Inventory export helpers ──────────────────────────────────────────────────

def export_inventory_excel(rows: list[dict], path: Path) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    data_row = _excel_header(ws, "MedicVista — Batch-wise Inventory Report", [])
    headers  = ["Product Name", "Company", "Packing",
                "Batch No", "Expiry", "Stock Qty", "Free Qty", "MRP", "Stock Value"]
    _excel_item_headers(ws, headers, data_row, fill_color="0f172a")
    data_row += 1
    total_val = 0.0
    for i, r in enumerate(rows):
        val = _safe_float(r.get("value", 0))
        total_val += val
        
        stock_val = r.get("stock", r.get("current_stock", 0))
        try:
            stock_val = float(str(stock_val).replace("₹", "").replace(",", "").strip())
        except ValueError:
            stock_val = _s(stock_val)
            
        free_val = r.get("current_free_qty", 0)
        try:
            free_val = float(str(free_val).replace("₹", "").replace(",", "").strip())
        except ValueError:
            free_val = _s(free_val)
            
        mrp_val = r.get("mrp", 0)
        try:
            mrp_val = float(str(mrp_val).replace("₹", "").replace(",", "").strip())
        except ValueError:
            mrp_val = _s(mrp_val)

        _excel_data_row(ws, data_row, [
            _s(r.get("name", r.get("product_name", ""))),
            _s(r.get("company", "")),
            _s(r.get("packing", "")),
            _s(r.get("batch_no", "")),
            _s(r.get("expiry", "")),
            stock_val,
            free_val,
            mrp_val,
            val,
        ], alt=(i % 2 == 0))
        data_row += 1
    _excel_totals_row(ws, data_row,
                       ["", "", "", "", "", "", "", "TOTAL", total_val])
    _auto_width(ws)
    wb.save(str(path))
    return str(path)


def export_inventory_pdf(rows: list[dict], path: Path, title="Batch-wise Inventory Report") -> str:
    W, H  = A4
    doc   = SimpleDocTemplate(str(path), pagesize=A4,
                               leftMargin=15*mm, rightMargin=15*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    usable = W - 30*mm
    story  = []
    banner = Table([[Paragraph(f"MedicVista — {title}", _H1)]], colWidths=[usable])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
    ]))
    story.append(banner)
    story.append(Spacer(1, 5*mm))

    col_w = [usable * x for x in [0.20, 0.12, 0.10, 0.12, 0.10,
                                    0.10, 0.08, 0.08, 0.10]]
    hdr = ["Product", "Company", "Packing", "Batch No", "Expiry",
           "Stock", "Free", "MRP", "Value"]
    tdata = [hdr]
    total_val = 0.0
    for r in rows:
        val = _safe_float(r.get("value", 0))
        total_val += val
        
        mrp_val = r.get("mrp", 0)
        try:
            mrp_val = float(str(mrp_val).replace("₹", "").replace(",", "").strip())
            mrp_str = _m(mrp_val)
        except ValueError:
            mrp_str = _s(mrp_val)

        tdata.append([
            _s(r.get("name", r.get("product_name", ""))),
            _s(r.get("company", "")),
            _s(r.get("packing", "")),
            _s(r.get("batch_no", "")),
            _s(r.get("expiry", "")),
            _s(r.get("stock", r.get("current_stock", ""))),
            _s(r.get("current_free_qty", "")),
            mrp_str,
            _m(val),
        ])
    tdata.append(["", "", "", "", "", "", "", "TOTAL", _m(total_val)])

    tbl = Table(tdata, colWidths=col_w, repeatRows=1)
    ts = _tbl_style()
    ts.add("FONTNAME",    (0, len(tdata) - 1), (-1, -1), "Helvetica-Bold")
    ts.add("BACKGROUND",  (0, len(tdata) - 1), (-1, -1), colors.HexColor("#DBEAFE"))
    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width=usable, thickness=0.5,
                              color=colors.HexColor("#E2E8F0")))
    story.append(Paragraph("Generated by MedicVista Pharmacy ERP", _SMALL))
    doc.build(story)
    return str(path)
