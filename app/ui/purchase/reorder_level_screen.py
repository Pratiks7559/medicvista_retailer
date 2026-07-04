"""Reorder Level Screen - PREMIUM EDITION
Shows products that need reordering based on sales patterns.
Glassmorphism-inspired panels, gradient accents, refined typography,
and micro-interactions throughout.
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, datetime, timedelta
from pathlib import Path

from ...styles import COLORS, FONT
from ...modern_ui_components import ModernButton, ModernEntry, ModernFrame, ModernLabel


# ----------------------------------------------------------------------------
# PREMIUM DESIGN TOKENS
# A self-contained palette so this screen feels like a distinct, elevated
# product surface rather than a default tkinter form.
# ----------------------------------------------------------------------------
PALETTE = {
    "bg_canvas":        "#0f1117",   # near-black canvas behind everything
    "bg_surface":       "#161927",   # card surface
    "bg_surface_alt":   "#1c2030",   # secondary surface (header strips, table head)
    "bg_surface_hover": "#20243a",
    "border":           "#2a2f45",
    "border_soft":      "#23273b",

    "text_primary":     "#f5f6fa",
    "text_secondary":   "#9aa1b9",
    "text_muted":       "#6b7290",

    "accent_a":         "#7c5cff",   # violet
    "accent_a_hover":   "#6c47ff",
    "accent_b":         "#22d3ee",   # cyan
    "accent_gradient_1":"#7c5cff",
    "accent_gradient_2":"#22d3ee",

    "danger":           "#ff5d6c",
    "danger_soft":      "#2a1620",
    "danger_soft_2":    "#321822",

    "success":          "#33e7a0",
    "success_soft":     "#102621",
    "success_soft_2":   "#0d2e26",

    "warning":          "#ffb454",

    "row_odd":          "#151824",
    "row_even":         "#181c2c",
}

FONT_FAMILY = "Segoe UI"


class ReorderLevelScreen(tk.Frame):
    def __init__(self, parent, app_instance, **kwargs):
        super().__init__(parent, bg=PALETTE["bg_canvas"], **kwargs)
        self.app = app_instance
        self._var_search = tk.StringVar()
        self._var_show_critical = tk.BooleanVar(value=False)
        self._rows: list[dict] = []

        self._summary_total = tk.StringVar(value="0")
        self._summary_critical = tk.StringVar(value="0")
        self._summary_sufficient = tk.StringVar(value="0")

        self._configure_treeview_style()
        self._build()
        self._bind_keys()
        self.on_show()

    # ------------------------------------------------------------------
    # STYLE CONFIGURATION
    # ------------------------------------------------------------------
    def _configure_treeview_style(self):
        """Dark, high-contrast treeview styling with generous row height."""
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Premium.Treeview",
            background=PALETTE["bg_surface"],
            fieldbackground=PALETTE["bg_surface"],
            foreground=PALETTE["text_primary"],
            rowheight=34,
            borderwidth=0,
            font=(FONT_FAMILY, 10),
        )
        style.configure(
            "Premium.Treeview.Heading",
            background=PALETTE["bg_surface_alt"],
            foreground=PALETTE["text_secondary"],
            font=(FONT_FAMILY, 10, "bold"),
            relief="flat",
            borderwidth=0,
        )
        style.map(
            "Premium.Treeview.Heading",
            background=[("active", PALETTE["bg_surface_hover"])],
        )
        style.map(
            "Premium.Treeview",
            background=[("selected", PALETTE["accent_a"])],
            foreground=[("selected", "#ffffff")],
        )
        style.layout(
            "Premium.Treeview",
            [("Premium.Treeview.treearea", {"sticky": "nswe"})],
        )

        style.configure(
            "Premium.Vertical.TScrollbar",
            background=PALETTE["bg_surface_alt"],
            troughcolor=PALETTE["bg_canvas"],
            bordercolor=PALETTE["bg_canvas"],
            arrowcolor=PALETTE["text_secondary"],
            relief="flat",
        )
        style.configure(
            "Premium.Horizontal.TScrollbar",
            background=PALETTE["bg_surface_alt"],
            troughcolor=PALETTE["bg_canvas"],
            bordercolor=PALETTE["bg_canvas"],
            arrowcolor=PALETTE["text_secondary"],
            relief="flat",
        )

    # ------------------------------------------------------------------
    # LAYOUT BUILD
    # ------------------------------------------------------------------
    def _build(self):
        main_container = tk.Frame(self, bg=PALETTE["bg_canvas"])
        main_container.pack(fill="both", expand=True, padx=28, pady=24)

        self._build_header(main_container)
        self.build_actions_section(main_container)
        self._build_summary_cards(main_container)
        self._build_tree_section(main_container)

    def _build_header(self, parent):
        header_frame = tk.Frame(parent, bg=PALETTE["bg_canvas"])
        header_frame.pack(fill="x", pady=(0, 22))

        header_left = tk.Frame(header_frame, bg=PALETTE["bg_canvas"])
        header_left.pack(side="left")

        # Thin gradient-style accent bar above the title
        accent_bar = tk.Frame(header_left, bg=PALETTE["accent_a"], height=3, width=46)
        accent_bar.pack(anchor="w", pady=(0, 10))

        title_label = tk.Label(
            header_left,
            text="Reorder Intelligence",
            font=(FONT_FAMILY, 26, "bold"),
            bg=PALETTE["bg_canvas"],
            fg=PALETTE["text_primary"],
        )
        title_label.pack(anchor="w")

        subtitle_label = tk.Label(
            header_left,
            text="Live stock signals, sales velocity & restocking priorities",
            font=(FONT_FAMILY, 11),
            bg=PALETTE["bg_canvas"],
            fg=PALETTE["text_secondary"],
        )
        subtitle_label.pack(anchor="w", pady=(4, 0))

        header_right = tk.Frame(header_frame, bg=PALETTE["bg_canvas"])
        header_right.pack(side="right", anchor="e")

        pill = tk.Frame(
            header_right, bg=PALETTE["success_soft"],
            highlightthickness=1, highlightbackground=PALETTE["success"],
        )
        pill.pack(side="right")
        pill_inner = tk.Frame(pill, bg=PALETTE["success_soft"], padx=14, pady=7)
        pill_inner.pack()

        self._dot = tk.Label(
            pill_inner, text="●", font=(FONT_FAMILY, 9),
            bg=PALETTE["success_soft"], fg=PALETTE["success"],
        )
        self._dot.pack(side="left", padx=(0, 6))

        self._quick_stats = tk.Label(
            pill_inner, text="System Ready",
            font=(FONT_FAMILY, 10, "bold"),
            bg=PALETTE["success_soft"], fg=PALETTE["success"],
        )
        self._quick_stats.pack(side="left")

    def build_actions_section(self, parent):
        """Glass-style toolbar with search, filter, refresh and export."""
        action_card = self._make_card(parent)
        action_card.pack(fill="x", pady=(0, 18))

        action_inner = tk.Frame(action_card, bg=PALETTE["bg_surface"], padx=22, pady=18)
        action_inner.pack(fill="x")

        filter_row = tk.Frame(action_inner, bg=PALETTE["bg_surface"])
        filter_row.pack(fill="x", pady=(0, 14))

        # --- Search box ------------------------------------------------
        search_wrap = tk.Frame(
            filter_row, bg=PALETTE["bg_surface_alt"],
            highlightthickness=1, highlightbackground=PALETTE["border"],
            highlightcolor=PALETTE["accent_a"],
        )
        search_wrap.pack(side="left", padx=(0, 14))

        search_inner = tk.Frame(search_wrap, bg=PALETTE["bg_surface_alt"])
        search_inner.pack(padx=12, pady=7)

        tk.Label(
            search_inner, text="⌕", font=(FONT_FAMILY, 13, "bold"),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["accent_b"],
        ).pack(side="left", padx=(0, 8))

        self._e_search = tk.Entry(
            search_inner,
            textvariable=self._var_search,
            font=(FONT_FAMILY, 11),
            bg=PALETTE["bg_surface_alt"],
            fg=PALETTE["text_primary"],
            insertbackground=PALETTE["accent_b"],
            relief="flat",
            borderwidth=0,
            width=32,
        )
        self._e_search.pack(side="left")
        self._e_search.config(highlightthickness=0)

        # focus glow effect
        self._e_search.bind("<FocusIn>", lambda e: search_wrap.config(highlightbackground=PALETTE["accent_a"]))
        self._e_search.bind("<FocusOut>", lambda e: search_wrap.config(highlightbackground=PALETTE["border"]))

        # --- Critical toggle --------------------------------------------
        toggle_wrap = tk.Frame(
            filter_row, bg=PALETTE["bg_surface_alt"],
            highlightthickness=1, highlightbackground=PALETTE["border"],
        )
        toggle_wrap.pack(side="left", padx=(0, 14))

        self._critical_check = tk.Checkbutton(
            toggle_wrap,
            text="  Critical Only",
            variable=self._var_show_critical,
            font=(FONT_FAMILY, 10, "bold"),
            bg=PALETTE["bg_surface_alt"],
            fg=PALETTE["danger"],
            activebackground=PALETTE["bg_surface_alt"],
            activeforeground=PALETTE["danger"],
            selectcolor=PALETTE["bg_surface_alt"],
            relief="flat",
            borderwidth=0,
            highlightthickness=0,
            command=self._fill_tree,
            cursor="hand2",
            padx=10, pady=7,
        )
        self._critical_check.pack()

        # --- Refresh button (gradient-style) -----------------------------
        refresh_btn = self._make_gradient_button(
            filter_row, "↻  Refresh", self.on_show,
            PALETTE["accent_a"], PALETTE["accent_a_hover"],
        )
        refresh_btn.pack(side="left")

        # --- Export button --------------------------------------------
        export_btn = self._make_gradient_button(
            filter_row, "⬇  Export Excel", self._export_excel,
            PALETTE["accent_b"], "#1bb8d4", text_color="#0a1f24",
        )
        export_btn.pack(side="right")

        # --- Hint row -----------------------------------------------------
        hint_row = tk.Frame(action_inner, bg=PALETTE["bg_surface"])
        hint_row.pack(fill="x", pady=(4, 0))

        formula_chip = tk.Frame(hint_row, bg=PALETTE["bg_surface_alt"])
        formula_chip.pack(side="left")
        tk.Label(
            formula_chip,
            text="ƒ  Reorder Level = (Avg Monthly Sales × 1.5) − Available Stock",
            font=(FONT_FAMILY, 9, "italic"),
            bg=PALETTE["bg_surface_alt"],
            fg=PALETTE["text_muted"],
            padx=10, pady=5,
        ).pack()

        legend_frame = tk.Frame(hint_row, bg=PALETTE["bg_surface"])
        legend_frame.pack(side="right")

        self._legend_item(legend_frame, PALETTE["danger"], "Critical")
        self._legend_item(legend_frame, PALETTE["success"], "Sufficient", pad_left=18)

    def _legend_item(self, parent, color, text, pad_left=0):
        wrap = tk.Frame(parent, bg=PALETTE["bg_surface"])
        wrap.pack(side="left", padx=(pad_left, 0))
        tk.Label(wrap, text="●", font=(FONT_FAMILY, 10), bg=PALETTE["bg_surface"], fg=color).pack(side="left")
        tk.Label(
            wrap, text=text, font=(FONT_FAMILY, 9, "bold"),
            bg=PALETTE["bg_surface"], fg=PALETTE["text_secondary"],
        ).pack(side="left", padx=(4, 0))

    def _make_gradient_button(self, parent, text, command, color, hover_color, text_color="white"):
        btn = tk.Button(
            parent,
            text=text,
            font=(FONT_FAMILY, 10, "bold"),
            bg=color,
            fg=text_color,
            activebackground=hover_color,
            activeforeground=text_color,
            relief="flat",
            borderwidth=0,
            padx=18,
            pady=9,
            cursor="hand2",
            command=command,
        )
        btn.bind("<Enter>", lambda e: btn.config(bg=hover_color))
        btn.bind("<Leave>", lambda e: btn.config(bg=color))
        return btn

    def _make_card(self, parent):
        """Reusable elevated 'glass card' container."""
        card = tk.Frame(
            parent, bg=PALETTE["bg_surface"],
            highlightthickness=1,
            highlightbackground=PALETTE["border"],
            highlightcolor=PALETTE["border"],
        )
        return card

    # ------------------------------------------------------------------
    # SUMMARY CARDS
    # ------------------------------------------------------------------
    def _build_summary_cards(self, parent):
        summary_frame = tk.Frame(parent, bg=PALETTE["bg_canvas"])
        summary_frame.pack(fill="x", pady=(0, 18))

        for i in range(3):
            summary_frame.grid_columnconfigure(i, weight=1)

        card_configs = [
            ("TOTAL PRODUCTS", self._summary_total, PALETTE["accent_a"], "◆", 0),
            ("CRITICAL ITEMS", self._summary_critical, PALETTE["danger"], "▲", 1),
            ("SUFFICIENT STOCK", self._summary_sufficient, PALETTE["success"], "●", 2),
        ]

        for title, var, color, glyph, col in card_configs:
            self._create_premium_card(summary_frame, title, var, color, glyph, col)

    def _create_premium_card(self, parent, title, var, color, glyph, col):
        card = tk.Frame(
            parent, bg=PALETTE["bg_surface"],
            highlightthickness=1, highlightbackground=PALETTE["border"],
        )
        pad = (0, 8) if col == 0 else (8, 0) if col == 2 else (8, 8)
        card.grid(row=0, column=col, sticky="ew", padx=pad)

        inner = tk.Frame(card, bg=PALETTE["bg_surface"], padx=22, pady=18)
        inner.pack(fill="both", expand=True)

        header_row = tk.Frame(inner, bg=PALETTE["bg_surface"])
        header_row.pack(fill="x")

        tk.Label(
            header_row, text=title, font=(FONT_FAMILY, 10, "bold"),
            bg=PALETTE["bg_surface"], fg=PALETTE["text_muted"],
        ).pack(side="left")

        glyph_badge = tk.Frame(header_row, bg=color)
        glyph_badge.pack(side="right")
        tk.Label(
            glyph_badge, text=glyph, font=(FONT_FAMILY, 9, "bold"),
            bg=color, fg=PALETTE["bg_canvas"], padx=7, pady=2,
        ).pack()

        value_row = tk.Frame(inner, bg=PALETTE["bg_surface"])
        value_row.pack(fill="x", pady=(10, 0))

        tk.Label(
            value_row, textvariable=var, font=(FONT_FAMILY, 34, "bold"),
            bg=PALETTE["bg_surface"], fg=PALETTE["text_primary"],
        ).pack(side="left")

        # progress-style underline accent
        divider_track = tk.Frame(inner, bg=PALETTE["bg_surface_alt"], height=4)
        divider_track.pack(fill="x", pady=(14, 0))
        divider_fill = tk.Frame(divider_track, bg=color, height=4, width=60)
        divider_fill.place(x=0, y=0)

    # ------------------------------------------------------------------
    # TREE / TABLE
    # ------------------------------------------------------------------
    def _build_tree_section(self, parent):
        tree_container = tk.Frame(
            parent, bg=PALETTE["bg_surface"],
            highlightthickness=1, highlightbackground=PALETTE["border"],
        )
        tree_container.pack(fill="both", expand=True)

        tree_header = tk.Frame(tree_container, bg=PALETTE["bg_surface_alt"], padx=22, pady=14)
        tree_header.pack(fill="x")

        title_wrap = tk.Frame(tree_header, bg=PALETTE["bg_surface_alt"])
        title_wrap.pack(side="left")

        tk.Frame(title_wrap, bg=PALETTE["accent_b"], width=4, height=18).pack(side="left", padx=(0, 10))
        tk.Label(
            title_wrap, text="Product Inventory Analysis",
            font=(FONT_FAMILY, 13, "bold"),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["text_primary"],
        ).pack(side="left")

        self._record_count = tk.Label(
            tree_header, text="",
            font=(FONT_FAMILY, 9, "bold"),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["text_muted"],
        )
        self._record_count.pack(side="right")

        tree_frame = tk.Frame(tree_container, bg=PALETTE["bg_surface"], padx=18, pady=16)
        tree_frame.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", style="Premium.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", style="Premium.Horizontal.TScrollbar")

        self.tree = ttk.Treeview(
            tree_frame,
            columns=("#", "Product", "Company", "Packing", "Avg Sale", "Reorder Level",
                     "Available", "Free Qty", "Need Reorder", "Status"),
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode="browse",
            height=18,
            style="Premium.Treeview",
        )

        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        columns = {
            "#": (50, "center"),
            "Product": (230, "w"),
            "Company": (160, "w"),
            "Packing": (110, "center"),
            "Avg Sale": (100, "center"),
            "Reorder Level": (120, "center"),
            "Available": (100, "center"),
            "Free Qty": (90, "center"),
            "Need Reorder": (120, "center"),
            "Status": (130, "center"),
        }

        for col, (width, anchor) in columns.items():
            self.tree.heading(col, text=col.upper())
            self.tree.column(col, width=width, anchor=anchor, minwidth=40)

        # Color tags
        self.tree.tag_configure("critical_odd", background=PALETTE["danger_soft"], foreground=PALETTE["danger"])
        self.tree.tag_configure("critical_even", background=PALETTE["danger_soft_2"], foreground=PALETTE["danger"])
        self.tree.tag_configure("odd", background=PALETTE["row_odd"], foreground=PALETTE["success"])
        self.tree.tag_configure("even", background=PALETTE["row_even"], foreground=PALETTE["success"])

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._show_batch_details)

        self._build_status_bar(tree_container)

    def _build_status_bar(self, parent):
        status_bar = tk.Frame(
            parent, bg=PALETTE["bg_surface_alt"], padx=22, pady=11,
            highlightthickness=1, highlightbackground=PALETTE["border"],
        )
        status_bar.pack(fill="x", side="bottom")

        left_status = tk.Frame(status_bar, bg=PALETTE["bg_surface_alt"])
        left_status.pack(side="left")

        tk.Label(
            left_status, text="◈", font=(FONT_FAMILY, 11),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["accent_a"],
        ).pack(side="left", padx=(0, 8))

        self._status_var = tk.StringVar(value="Ready")
        tk.Label(
            left_status, textvariable=self._status_var, font=(FONT_FAMILY, 10),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["text_secondary"],
        ).pack(side="left")

        right_status = tk.Frame(status_bar, bg=PALETTE["bg_surface_alt"])
        right_status.pack(side="right")

        shortcuts = [("F5", "Refresh"), ("Enter", "Search"), ("Double-click", "Batch Details")]
        for i, (key, action) in enumerate(shortcuts):
            if i > 0:
                tk.Label(
                    right_status, text="·", font=(FONT_FAMILY, 10),
                    bg=PALETTE["bg_surface_alt"], fg=PALETTE["border"],
                ).pack(side="left", padx=10)

            key_chip = tk.Frame(right_status, bg=PALETTE["bg_surface"], highlightthickness=1,
                                 highlightbackground=PALETTE["border"])
            key_chip.pack(side="left")
            tk.Label(
                key_chip, text=key, font=(FONT_FAMILY, 8, "bold"),
                bg=PALETTE["bg_surface"], fg=PALETTE["accent_b"], padx=7, pady=2,
            ).pack()

            tk.Label(
                right_status, text=action, font=(FONT_FAMILY, 9),
                bg=PALETTE["bg_surface_alt"], fg=PALETTE["text_muted"],
            ).pack(side="left", padx=(6, 0))

    # ------------------------------------------------------------------
    # KEY BINDINGS
    # ------------------------------------------------------------------
    def _bind_keys(self):
        self._e_search.bind("<Return>", lambda e: self.on_show())
        self.tree.bind("<Return>", lambda e: self._show_batch_details())
        self.bind_all("<F5>", lambda e: self.on_show())

    def kb_refresh(self):
        self.on_show()

    def kb_search(self):
        self._e_search.focus_set()

    # ------------------------------------------------------------------
    # DATA LOADING
    # ------------------------------------------------------------------
    def on_show(self):
        """Load reorder level data."""
        self._rows = []
        for i in self.tree.get_children():
            self.tree.delete(i)

        try:
            self._status_var.set("Loading reorder data…")
            self.update_idletasks()

            reorder_data = self._calculate_reorder_levels()
            self._rows = reorder_data

            self._summary_total.set(str(len(reorder_data)))
            critical = sum(1 for r in reorder_data if r["status"] == "critical")
            self._summary_critical.set(str(critical))
            self._summary_sufficient.set(str(len(reorder_data) - critical))

            self._fill_tree()
            self._status_var.set(f"Loaded {len(reorder_data)} products")

            self._quick_stats.config(text=f"{len(reorder_data)} Products Loaded")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load reorder data: {e}")
            self._status_var.set("Error loading data")
            self._quick_stats.config(text="Error", fg=PALETTE["danger"])

    def _calculate_reorder_levels(self) -> list[dict]:
        """Calculate reorder levels based on sales and stock."""
        try:
            thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            print(f"[DEBUG] Calculating reorder from date: {thirty_days_ago}")

            products = self.app.db.query("""
                SELECT productid, product_name, product_company, product_packing
                FROM core_productmaster
                WHERE retailer_id = %s
                ORDER BY product_name
                LIMIT 100
            """, (self.app.db.config.retailer_id,))

            print(f"[DEBUG] Found {len(products)} products in database")

            if not products:
                print("[DEBUG] No products found in database")
                return []

            reorder_data = []

            for product in products:
                product_id = product["productid"]

                stock_result = self.app.db.query("""
                    SELECT 
                        COALESCE(SUM(CASE
                            WHEN transaction_type IN ('PURCHASE') THEN quantity + free_quantity
                            WHEN transaction_type IN ('PURCHASE_RETURN','REVERT_PURCHASE_RETURN','REVERT_PURCHASE') 
                                THEN -(quantity + free_quantity)
                            WHEN transaction_type IN ('SALE','STOCK_ISSUE','REVERT_SALE') 
                                THEN -(quantity + free_quantity)
                            WHEN transaction_type IN ('SALES_RETURN','REVERT_SALES_RETURN') 
                                THEN (quantity + free_quantity)
                            ELSE 0 END), 0) AS total_stock,
                        COALESCE(SUM(CASE
                            WHEN transaction_type IN ('PURCHASE') THEN free_quantity
                            WHEN transaction_type IN ('PURCHASE_RETURN','REVERT_PURCHASE_RETURN','REVERT_PURCHASE') 
                                THEN -free_quantity
                            WHEN transaction_type IN ('SALE','STOCK_ISSUE','REVERT_SALE') 
                                THEN -free_quantity
                            WHEN transaction_type IN ('SALES_RETURN','REVERT_SALES_RETURN') 
                                THEN free_quantity
                            ELSE 0 END), 0) AS free_qty
                    FROM inventory_transaction
                    WHERE product_id = %s AND retailer_id = %s
                """, (product_id, self.app.db.config.retailer_id))

                total_stock = float(stock_result[0]["total_stock"]) if stock_result else 0
                free_qty = float(stock_result[0]["free_qty"]) if stock_result else 0

                sales_result = self.app.db.query("""
                    SELECT COALESCE(SUM(quantity), 0) AS total_sales
                    FROM inventory_transaction
                    WHERE product_id = %s 
                        AND retailer_id = %s
                        AND transaction_type = 'SALE'
                        AND DATE(transaction_date) >= %s
                """, (product_id, self.app.db.config.retailer_id, thirty_days_ago))

                avg_monthly_sale = float(sales_result[0]["total_sales"]) if sales_result else 0
                reorder_level = avg_monthly_sale * 1.5
                reorder_needed = max(0, reorder_level - total_stock)
                status = "critical" if reorder_needed > 0 else "sufficient"

                reorder_data.append({
                    "product_id": product_id,
                    "product_name": product["product_name"],
                    "product_company": product["product_company"],
                    "product_packing": product["product_packing"],
                    "avg_monthly_sale": avg_monthly_sale,
                    "reorder_level": reorder_level,
                    "total_available": total_stock,
                    "free_qty": free_qty,
                    "reorder_needed": reorder_needed,
                    "status": status,
                })

            print(f"[DEBUG] Calculated reorder levels for {len(reorder_data)} products")
            reorder_data.sort(key=lambda x: x["reorder_needed"], reverse=True)

            return reorder_data

        except Exception as e:
            print(f"[DEBUG ERROR] Failed to calculate reorder levels: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _fill_tree(self):
        """Fill tree with filtered data."""
        term = self._var_search.get().lower()
        show_critical = self._var_show_critical.get()

        for i in self.tree.get_children():
            self.tree.delete(i)

        shown = 0
        for idx, r in enumerate(self._rows, 1):
            if term and term not in r["product_name"].lower() and term not in r["product_company"].lower():
                continue

            if show_critical and r["status"] != "critical":
                continue

            if r["status"] == "critical":
                row_tag = "critical_odd" if idx % 2 else "critical_even"
            else:
                row_tag = "odd" if idx % 2 else "even"

            status_display = "▲ CRITICAL" if r["status"] == "critical" else "● SUFFICIENT"

            self.tree.insert("", "end", iid=str(r["product_id"]),
                              tags=(row_tag,), values=(
                idx,
                r["product_name"],
                r["product_company"],
                r["product_packing"],
                f"{r['avg_monthly_sale']:.1f}",
                f"{r['reorder_level']:.1f}",
                f"{r['total_available']:.1f}",
                f"{r['free_qty']:.1f}",
                f"{r['reorder_needed']:.1f}",
                status_display,
            ))
            shown += 1

        self._status_var.set(f"Showing {shown} / {len(self._rows)} products")
        self._record_count.config(text=f"{shown} RECORDS")

    def _show_batch_details(self, event=None):
        """Show batch-wise stock details for selected product."""
        sel = self.tree.focus()
        if not sel:
            return

        product_id = int(sel)

        product = next((r for r in self._rows if r["product_id"] == product_id), None)
        if not product:
            return

        batches = self.app.db.query("""
            SELECT 
                t.batch_no,
                t.expiry_date,
                COALESCE(MAX(t.mrp), 0) AS mrp,
                COALESCE(MAX(t.rate), 0) AS purchase_rate,
                COALESCE(SUM(CASE
                    WHEN t.transaction_type IN ('PURCHASE') THEN t.quantity
                    WHEN t.transaction_type IN ('PURCHASE_RETURN','REVERT_PURCHASE_RETURN','REVERT_PURCHASE') 
                        THEN -t.quantity
                    WHEN t.transaction_type IN ('SALE','STOCK_ISSUE','REVERT_SALE') 
                        THEN -t.quantity
                    WHEN t.transaction_type IN ('SALES_RETURN','REVERT_SALES_RETURN') 
                        THEN t.quantity
                    ELSE 0 END), 0) AS stock,
                COALESCE(SUM(CASE
                    WHEN t.transaction_type IN ('PURCHASE') THEN t.free_quantity
                    WHEN t.transaction_type IN ('PURCHASE_RETURN','REVERT_PURCHASE_RETURN','REVERT_PURCHASE') 
                        THEN -t.free_quantity
                    WHEN t.transaction_type IN ('SALE','STOCK_ISSUE','REVERT_SALE') 
                        THEN -t.free_quantity
                    WHEN t.transaction_type IN ('SALES_RETURN','REVERT_SALES_RETURN') 
                        THEN t.free_quantity
                    ELSE 0 END), 0) AS free_qty
            FROM inventory_transaction t
            WHERE t.product_id = %s AND t.retailer_id = %s
            GROUP BY t.batch_no, t.expiry_date
            HAVING stock > 0 OR free_qty > 0
            ORDER BY t.expiry_date
        """, (product_id, self.app.db.config.retailer_id))

        BatchDetailsDialog(self, product, batches)

    # ------------------------------------------------------------------
    # EXCEL EXPORT (premium dark-themed workbook)
    # ------------------------------------------------------------------
    def _export_excel(self):
        """Export reorder level report to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            if not self._rows:
                messagebox.showwarning("No Data", "No data to export.", parent=self)
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Reorder Level"

            ws['A1'] = "REORDER LEVEL REPORT"
            ws['A1'].font = Font(size=18, bold=True, color="1a1c2e")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:J1')

            ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
            ws['A2'].font = Font(size=10, color="6b7290")
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:J2')

            headers = ["#", "Product Name", "Company", "Packing", "Avg Monthly Sale",
                       "Reorder Level", "Available Stock", "Free Qty", "Reorder Needed", "Status"]

            header_fill = PatternFill(start_color='7c5cff', end_color='7c5cff', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            border = Border(left=Side(style='thin', color='2a2f45'),
                             right=Side(style='thin', color='2a2f45'),
                             top=Side(style='thin', color='2a2f45'),
                             bottom=Side(style='thin', color='2a2f45'))

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, row in enumerate(self._rows, 1):
                row_num = idx + 4

                if row["status"] == "critical":
                    fill = PatternFill(start_color='2a1620', end_color='2a1620', fill_type='solid')
                    font = Font(color='ff5d6c', bold=True)
                else:
                    fill = PatternFill(start_color='102621', end_color='102621', fill_type='solid')
                    font = Font(color='33e7a0')

                ws.cell(row=row_num, column=1, value=idx).font = font
                ws.cell(row=row_num, column=2, value=row["product_name"]).font = font
                ws.cell(row=row_num, column=3, value=row["product_company"]).font = font
                ws.cell(row=row_num, column=4, value=row["product_packing"]).font = font
                ws.cell(row=row_num, column=5, value=round(row["avg_monthly_sale"], 1)).font = font
                ws.cell(row=row_num, column=6, value=round(row["reorder_level"], 1)).font = font
                ws.cell(row=row_num, column=7, value=round(row["total_available"], 1)).font = font
                ws.cell(row=row_num, column=8, value=round(row["free_qty"], 1)).font = font
                ws.cell(row=row_num, column=9, value=round(row["reorder_needed"], 1)).font = font
                ws.cell(row=row_num, column=10, value=row["status"].upper()).font = font

                for col in range(1, 11):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal='center' if col != 2 and col != 3 else 'left',
                        vertical='center',
                    )

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 18
            ws.column_dimensions['J'].width = 12

            docs_dir = Path.home() / "Documents" / "MedicVista_Exports"
            docs_dir.mkdir(parents=True, exist_ok=True)

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=f"Reorder_Level_{date.today().isoformat()}.xlsx",
                initialdir=str(docs_dir),
                title="Save Reorder Level Report",
                parent=self,
            )

            if filename:
                wb.save(filename)
                if messagebox.askyesno("Export Complete",
                                        f"Saved to:\n{filename}\n\nOpen file now?", parent=self):
                    import os
                    os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}", parent=self)


class BatchDetailsDialog(tk.Toplevel):
    """Premium dark dialog showing batch-wise stock details."""

    def __init__(self, parent, product, batches):
        super().__init__(parent)
        self.product = product
        self.batches = batches

        self.title(f"Batch Details — {product['product_name']}")
        self.geometry("920x560")
        self.configure(bg=PALETTE["bg_canvas"])
        self.grab_set()
        self._configure_local_style()
        self._center_window()
        self._build()

    def _configure_local_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(
            "BatchDialog.Treeview",
            background=PALETTE["bg_surface"],
            fieldbackground=PALETTE["bg_surface"],
            foreground=PALETTE["text_primary"],
            rowheight=32,
            borderwidth=0,
            font=(FONT_FAMILY, 10),
        )
        style.configure(
            "BatchDialog.Treeview.Heading",
            background=PALETTE["bg_surface_alt"],
            foreground=PALETTE["text_secondary"],
            font=(FONT_FAMILY, 10, "bold"),
            relief="flat",
        )
        style.map(
            "BatchDialog.Treeview",
            background=[("selected", PALETTE["accent_a"])],
            foreground=[("selected", "#ffffff")],
        )
        style.configure(
            "BatchDialog.Vertical.TScrollbar",
            background=PALETTE["bg_surface_alt"],
            troughcolor=PALETTE["bg_canvas"],
            arrowcolor=PALETTE["text_secondary"],
            relief="flat",
        )

    def _center_window(self):
        self.update_idletasks()
        width, height = 920, 560
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def _build(self):
        # ---- Header with violet-cyan gradient feel ----------------------
        header = tk.Frame(self, bg=PALETTE["accent_a"], pady=18, padx=26)
        header.pack(fill="x")

        accent_strip = tk.Frame(self, bg=PALETTE["accent_b"], height=3)
        accent_strip.pack(fill="x")

        name_frame = tk.Frame(header, bg=PALETTE["accent_a"])
        name_frame.pack(fill="x")

        tk.Label(
            name_frame, text="◆", font=(FONT_FAMILY, 16),
            bg=PALETTE["accent_a"], fg="white",
        ).pack(side="left", padx=(0, 10))

        tk.Label(
            name_frame, text=self.product["product_name"],
            font=(FONT_FAMILY, 17, "bold"), bg=PALETTE["accent_a"], fg="white",
        ).pack(side="left")

        status_color = PALETTE["danger"] if self.product["status"] == "critical" else PALETTE["success"]
        status_text = "▲ CRITICAL" if self.product["status"] == "critical" else "● SUFFICIENT"

        status_chip = tk.Frame(header, bg=status_color)
        status_chip.pack(side="right", anchor="n")
        tk.Label(
            status_chip, text=status_text, font=(FONT_FAMILY, 10, "bold"),
            bg=status_color, fg="white" if status_color == PALETTE["danger"] else "#0a1f1a",
            padx=14, pady=6,
        ).pack()

        details_frame = tk.Frame(header, bg=PALETTE["accent_a"])
        details_frame.pack(fill="x", pady=(8, 0))

        details = [
            f"🏢  {self.product['product_company']}",
            f"📦  {self.product['product_packing']}",
            f"📈  Avg Sale: {self.product['avg_monthly_sale']:.1f}",
            f"🔄  Reorder Level: {self.product['reorder_level']:.1f}",
        ]
        for d in details:
            tk.Label(
                details_frame, text=d, font=(FONT_FAMILY, 10),
                bg=PALETTE["accent_a"], fg="#e3d9ff",
            ).pack(side="left", padx=(0, 18))

        # ---- Summary stat strip ------------------------------------------
        summary = tk.Frame(
            self, bg=PALETTE["bg_surface"], padx=22, pady=18,
            highlightthickness=1, highlightbackground=PALETTE["border"],
        )
        summary.pack(fill="x", padx=20, pady=20)

        total_stock = self.product["total_available"]
        total_free = self.product["free_qty"]
        need_reorder = self.product["reorder_needed"]

        stats = [
            (f"{total_stock:.1f}", "Total Stock", PALETTE["text_primary"]),
            (f"{total_free:.1f}", "Free Qty", PALETTE["accent_b"]),
            (f"{need_reorder:.1f}", "Need Reorder", PALETTE["danger"] if need_reorder > 0 else PALETTE["success"]),
        ]

        stat_frame = tk.Frame(summary, bg=PALETTE["bg_surface"])
        stat_frame.pack(fill="x")

        for i, (value, label, color) in enumerate(stats):
            if i > 0:
                tk.Frame(stat_frame, bg=PALETTE["border"], width=1).pack(side="left", fill="y", padx=20)

            group = tk.Frame(stat_frame, bg=PALETTE["bg_surface"])
            group.pack(side="left")

            tk.Label(
                group, text=value, font=(FONT_FAMILY, 18, "bold"),
                bg=PALETTE["bg_surface"], fg=color,
            ).pack()
            tk.Label(
                group, text=label.upper(), font=(FONT_FAMILY, 8, "bold"),
                bg=PALETTE["bg_surface"], fg=PALETTE["text_muted"],
            ).pack(pady=(2, 0))

        # ---- Batch table -------------------------------------------------
        table_frame = tk.Frame(self, bg=PALETTE["bg_canvas"], padx=20)
        table_frame.pack(fill="both", expand=True, pady=(0, 20))

        table_header = tk.Frame(table_frame, bg=PALETTE["bg_canvas"])
        table_header.pack(fill="x", pady=(0, 10))

        tk.Label(
            table_header, text="Batch-wise Stock Details",
            font=(FONT_FAMILY, 13, "bold"), bg=PALETTE["bg_canvas"], fg=PALETTE["text_primary"],
        ).pack(side="left")

        tk.Label(
            table_header, text=f"{len(self.batches)} BATCHES",
            font=(FONT_FAMILY, 9, "bold"), bg=PALETTE["bg_canvas"], fg=PALETTE["text_muted"],
        ).pack(side="right")

        frm = tk.Frame(table_frame, bg=PALETTE["bg_surface"], highlightthickness=1,
                        highlightbackground=PALETTE["border"])
        frm.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(frm, orient="vertical", style="BatchDialog.Vertical.TScrollbar")
        tree = ttk.Treeview(
            frm,
            columns=("Batch", "Expiry", "MRP", "Rate", "Stock", "Free Qty", "Total"),
            show="headings",
            yscrollcommand=vsb.set,
            selectmode="browse",
            height=12,
            style="BatchDialog.Treeview",
        )
        vsb.config(command=tree.yview)

        columns = {
            "Batch": (160, "w"), "Expiry": (110, "center"), "MRP": (100, "center"),
            "Rate": (100, "center"), "Stock": (100, "center"),
            "Free Qty": (100, "center"), "Total": (110, "center"),
        }
        for col, (width, anchor) in columns.items():
            tree.heading(col, text=col.upper())
            tree.column(col, width=width, anchor=anchor)

        tree.tag_configure("odd", background=PALETTE["row_odd"])
        tree.tag_configure("even", background=PALETTE["row_even"])

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        frm.grid_rowconfigure(0, weight=1)
        frm.grid_columnconfigure(0, weight=1)

        for idx, batch in enumerate(self.batches):
            stock = float(batch["stock"])
            free = float(batch["free_qty"])
            total = stock + free
            tag = "odd" if idx % 2 else "even"

            tree.insert("", "end", tags=(tag,), values=(
                batch["batch_no"],
                batch["expiry_date"],
                f"₹{batch['mrp']:.2f}",
                f"₹{batch['purchase_rate']:.2f}",
                f"{stock:.1f}",
                f"{free:.1f}",
                f"{total:.1f}",
            ))

        # ---- Close button --------------------------------------------
        btn_frame = tk.Frame(self, bg=PALETTE["bg_canvas"], pady=16)
        btn_frame.pack(fill="x")

        close_btn = tk.Button(
            btn_frame, text="Close", font=(FONT_FAMILY, 10, "bold"),
            bg=PALETTE["bg_surface_alt"], fg=PALETTE["text_primary"],
            activebackground=PALETTE["bg_surface_hover"], activeforeground=PALETTE["text_primary"],
            relief="flat", borderwidth=0, padx=28, pady=10, cursor="hand2",
            highlightthickness=1, highlightbackground=PALETTE["border"],
            command=self.destroy,
        )
        close_btn.pack()
        close_btn.bind("<Enter>", lambda e: close_btn.config(bg=PALETTE["bg_surface_hover"]))
        close_btn.bind("<Leave>", lambda e: close_btn.config(bg=PALETTE["bg_surface_alt"]))