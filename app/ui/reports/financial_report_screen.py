import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date
from ...styles import COLORS, FONT

BG = COLORS["bg_light"]
CARD = "#ffffff"
DARK = "#1e293b"
DARKER = "#0f172a"
BDR = "#334155"
TXT_LT = "#f1f5f9"
TXT_DK = "#1e293b"
MUTED = "#94a3b8"
BLUE = "#3b82f6"
BLUE_H = "#2563eb"
GREEN = "#10b981"
GREEN_H = "#059669"
RED = "#ef4444"
RED_H = "#dc2626"
PURPLE = "#7c3aed"

class FinancialReportScreen(tk.Frame):
    def __init__(self, parent, app_instance, **kwargs):
        super().__init__(parent, bg=BG, **kwargs)
        self.app = app_instance
        self._var_from = tk.StringVar()
        self._var_to = tk.StringVar()
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=DARK, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="💰 Financial Report - Profit Analysis", font=("Segoe UI", 15, "bold"),
                 fg=TXT_LT, bg=DARK).pack(side="left")
        
        self._btn(hdr, "📊 Excel", GREEN, GREEN_H, self._export_excel).pack(side="right", padx=3)
        self._btn(hdr, "📄 PDF", RED, RED_H, self._export_pdf).pack(side="right", padx=3)

        flt = tk.Frame(self, bg=DARK, padx=20, pady=14)
        flt.pack(fill="x")
        from ..purchase.purchase_invoice_dialog import DateEntry as DE
        self._var_from.set(date.today().replace(day=1).isoformat())
        self._var_to.set(date.today().isoformat())
        
        tk.Label(flt, text="From", font=("Segoe UI", 9), fg=MUTED, bg=DARK).grid(row=0, column=0, sticky="w")
        DE(flt, textvariable=self._var_from, width=14, bg=DARK).grid(row=1, column=0, padx=(0,20))
        tk.Label(flt, text="To", font=("Segoe UI", 9), fg=MUTED, bg=DARK).grid(row=0, column=1, sticky="w")
        DE(flt, textvariable=self._var_to, width=14, bg=DARK).grid(row=1, column=1, padx=(0,16))
        self._btn(flt, "🔍 Show", BLUE, BLUE_H, self._load).grid(row=1, column=2)

        cf = tk.Frame(self, bg=BG, padx=16, pady=14)
        cf.pack(fill="x")
        self._v_sales = tk.StringVar(value="₹0")
        self._v_purchase = tk.StringVar(value="₹0")
        self._v_profit = tk.StringVar(value="₹0")
        self._v_gst = tk.StringVar(value="₹0")
        
        stat_items = [("Sales", self._v_sales, "#7c3aed"), ("Purchase", self._v_purchase, "#ec4899"),
                      ("Profit", self._v_profit, "#10b981"), ("GST", self._v_gst, "#f59e0b")]
        
        for idx, (lbl, var, bg) in enumerate(stat_items):
            c = tk.Frame(cf, bg=bg, padx=20, pady=16)
            c.pack(side="left", expand=True, fill="x", padx=5)
            tk.Label(c, text=lbl, font=("Segoe UI", 9), fg="white", bg=bg).pack(anchor="w")
            tk.Label(c, textvariable=var, font=("Segoe UI", 18, "bold"), fg="white", bg=bg).pack(anchor="w")

        tf = tk.Frame(self, bg=DARK, padx=16, pady=0)
        tf.pack(fill="both", expand=True)
        
        cols = ("Date", "Type", "Invoice", "Party", "Product", "Batch", "Qty", "P.Rate", "S.Rate", "GST", "P.Cost", "S.Value", "Profit")
        widths = (90, 100, 120, 150, 180, 80, 60, 80, 80, 70, 90, 90, 90)
        
        style = ttk.Style()
        style.configure("Dark.Treeview", rowheight=34, font=("Segoe UI", 9), background=DARK,
                        fieldbackground=DARK, foreground=TXT_LT, bordercolor=BDR)
        style.configure("Dark.Treeview.Heading", font=("Segoe UI", 9, "bold"), background=DARKER,
                        foreground=MUTED, relief="flat")
        style.map("Dark.Treeview", background=[("selected", BLUE)], foreground=[("selected", "white")])
        
        vsb = ttk.Scrollbar(tf, orient="vertical")
        hsb = ttk.Scrollbar(tf, orient="horizontal")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", style="Dark.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set, selectmode="browse")
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col, anchor="w")
            self.tree.column(col, width=w, anchor="w" if col in ("Party","Product") else "center")
        
        self.tree.tag_configure("total", font=("Segoe UI", 9, "bold"), background=DARKER, foreground=TXT_LT)
        self.tree.tag_configure("profit", foreground=GREEN)
        self.tree.tag_configure("loss", foreground=RED)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

    def _btn(self, parent, text, bg, hover, cmd):
        b = tk.Button(parent, text=text, bg=bg, fg="white", font=("Segoe UI", 9, "bold"),
                      bd=0, relief="flat", padx=14, pady=6, cursor="hand2", command=cmd,
                      activebackground=hover, activeforeground="white")
        b.bind("<Enter>", lambda e: b.config(bg=hover))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    def _load(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        frm, to = self._var_from.get().strip(), self._var_to.get().strip()
        
        try:
            rows, summary = self._build_financial_data(frm, to)
            self._v_sales.set(f"₹{summary['sales']:,.2f}")
            self._v_purchase.set(f"₹{summary['purchase']:,.2f}")
            self._v_profit.set(f"₹{summary['profit']:,.2f}")
            self._v_gst.set(f"₹{summary['gst']:,.2f}")
            
            for r in rows:
                tag = "profit" if r['profit'] >= 0 else "loss"
                self.tree.insert("", "end", values=(
                    str(r['date'])[:10], r['type'], r['invoice'], r['party'][:20],
                    r['product'][:25], r['batch'][:10], f"{r['qty']:.0f}",
                    f"₹{r['p_rate']:.2f}", f"₹{r['s_rate']:.2f}", f"₹{r['gst']:.2f}",
                    f"₹{r['p_cost']:.2f}", f"₹{r['s_value']:.2f}", f"₹{r['profit']:.2f}"
                ), tags=(tag,))
            
            self.tree.insert("", "end", values=("", "TOTAL", "", "", "", "", "",
                "", "", f"₹{summary['gst']:,.2f}", f"₹{summary['purchase']:,.2f}",
                f"₹{summary['sales']:,.2f}", f"₹{summary['profit']:,.2f}"), tags=("total",))
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {e}", parent=self)

    def _build_financial_data(self, frm, to):
        rows = []
        t_sales = t_purchase = t_gst = t_profit = 0.0
        
        # Sales
        q = "SELECT s.*, si.sales_invoice_date, c.customer_name FROM core_salesmaster s JOIN core_salesinvoicemaster si ON s.sales_invoice_no_id=si.sales_invoice_no JOIN core_customermaster c ON s.customerid_id=c.customerid WHERE 1=1"
        p = []
        if frm: q += " AND si.sales_invoice_date >= %s"; p.append(frm)
        if to: q += " AND si.sales_invoice_date <= %s"; p.append(to)
        q += " ORDER BY si.sales_invoice_date DESC LIMIT 500"
        
        lookup = self._build_purchase_lookup()
        for s in self.app.db.query(q, p):
            pr = lookup.get((s.get('productid_id'), s.get('product_batch_no')), 0.0)
            qty = float(s.get('sale_quantity', 0))
            sr = float(s.get('sale_rate', 0))
            cgst = float(s.get('sale_cgst', 0))
            sgst = float(s.get('sale_sgst', 0))
            sv = sr * qty
            pc = pr * qty
            gst = sv * (cgst + sgst) / 100
            pft = sv - pc
            rows.append({'date': s.get('sales_invoice_date'), 'type': 'Sale', 'invoice': s.get('sales_invoice_no'),
                         'party': s.get('customer_name', ''), 'product': s.get('product_name', ''),
                         'batch': s.get('product_batch_no', ''), 'qty': qty, 'p_rate': pr, 's_rate': sr,
                         'gst': gst, 'p_cost': pc, 's_value': sv, 'profit': pft})
            t_sales += sv; t_purchase += pc; t_gst += gst; t_profit += pft
        
        # Purchases
        q = "SELECT p.*, i.invoice_date, s.supplier_name FROM core_purchasemaster p JOIN core_invoicemaster i ON p.product_invoiceid_id=i.invoiceid JOIN core_suppliermaster s ON p.product_supplierid_id=s.supplierid WHERE 1=1"
        p = []
        if frm: q += " AND i.invoice_date >= %s"; p.append(frm)
        if to: q += " AND i.invoice_date <= %s"; p.append(to)
        q += " ORDER BY i.invoice_date DESC LIMIT 500"
        
        for pur in self.app.db.query(q, p):
            qty = float(pur.get('product_quantity', 0))
            pr = float(pur.get('product_purchase_rate', 0))
            cgst = float(pur.get('CGST', 0))
            sgst = float(pur.get('SGST', 0))
            pc = pr * qty
            gst = pc * (cgst + sgst) / 100
            rows.append({'date': pur.get('invoice_date'), 'type': 'Purchase', 'invoice': pur.get('product_invoice_no'),
                         'party': pur.get('supplier_name', ''), 'product': pur.get('product_name', ''),
                         'batch': pur.get('product_batch_no', ''), 'qty': qty, 'p_rate': pr, 's_rate': 0,
                         'gst': gst, 'p_cost': pc, 's_value': 0, 'profit': -pc})
            t_purchase += pc; t_gst += gst; t_profit -= pc
        
        # Sales Returns
        q = "SELECT r.*, ri.return_sales_invoice_date, c.customer_name FROM core_returnsalesmaster r JOIN core_returnsalesinvoicemaster ri ON r.return_sales_invoice_no_id=ri.return_sales_invoice_no JOIN core_customermaster c ON r.return_customerid_id=c.customerid WHERE 1=1"
        p = []
        if frm: q += " AND ri.return_sales_invoice_date >= %s"; p.append(frm)
        if to: q += " AND ri.return_sales_invoice_date <= %s"; p.append(to)
        q += " ORDER BY ri.return_sales_invoice_date DESC LIMIT 500"
        
        for r in self.app.db.query(q, p):
            pr = lookup.get((r.get('return_productid_id'), r.get('return_product_batch_no')), 0.0)
            qty = float(r.get('return_sale_quantity', 0))
            sr = float(r.get('return_sale_rate', 0))
            cgst = float(r.get('return_sale_cgst', 0))
            sgst = float(r.get('return_sale_sgst', 0))
            sv = sr * qty
            pc = pr * qty
            gst = sv * (cgst + sgst) / 100
            pft = -(sv - pc)
            rows.append({'date': r.get('return_sales_invoice_date'), 'type': 'Sale Return', 'invoice': r.get('return_sales_invoice_no'),
                         'party': r.get('customer_name', ''), 'product': r.get('return_product_name', ''),
                         'batch': r.get('return_product_batch_no', ''), 'qty': -qty, 'p_rate': pr, 's_rate': sr,
                         'gst': -gst, 'p_cost': -pc, 's_value': -sv, 'profit': pft})
            t_sales -= sv; t_purchase -= pc; t_gst -= gst; t_profit += pft
        
        rows.sort(key=lambda x: x['date'], reverse=True)
        return rows, {'sales': t_sales, 'purchase': t_purchase, 'gst': t_gst, 'profit': t_profit}

    def _build_purchase_lookup(self):
        lookup = {}
        for p in self.app.db.query("SELECT productid_id, product_batch_no, product_purchase_rate FROM core_purchasemaster", []):
            lookup[(p.get('productid_id'), p.get('product_batch_no'))] = float(p.get('product_purchase_rate', 0))
        return lookup

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                                              initialfile=f"financial_report_{date.today()}.xlsx", parent=self)
            if not fp: return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Report"
            ws['A1'] = f"Financial Report | {self._var_from.get()} to {self._var_to.get()}"
            ws['A1'].font = Font(bold=True, size=12)
            
            hdrs = ["Date", "Type", "Invoice", "Party", "Product", "Batch", "Qty", "P.Rate", "S.Rate", "GST", "P.Cost", "S.Value", "Profit"]
            for i, h in enumerate(hdrs, 1):
                c = ws.cell(3, i, h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            
            row = 4
            for item in self.tree.get_children():
                if "total" not in self.tree.item(item, "tags"):
                    vals = self.tree.item(item, "values")
                    for col, v in enumerate(vals, 1):
                        ws.cell(row, col, str(v).replace("₹","").replace(",","").strip())
                    row += 1
            
            wb.save(fp)
            messagebox.showinfo("Success", f"Saved to:\n{fp}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}", parent=self)

    def _export_pdf(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            from pathlib import Path
            
            # Get save location
            docs_dir = Path.home() / "Documents" / "MedicVista_Exports"
            docs_dir.mkdir(parents=True, exist_ok=True)
            
            fp = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Files", "*.pdf")],
                initialfile=f"financial_report_{date.today()}.pdf",
                initialdir=str(docs_dir),
                parent=self
            )
            if not fp:
                return
            
            # Create PDF document (landscape for wide table)
            doc = SimpleDocTemplate(fp, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a237e'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("💰 Financial Report - Profit Analysis", title_style))
            
            # Date range
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(f"Period: {self._var_from.get()} to {self._var_to.get()}", date_style))
            
            # Summary stats
            summary_data = [
                ['Total Sales', 'Total Purchase', 'GST Collected', 'Net Profit'],
                [self._v_sales.get(), self._v_purchase.get(), self._v_gst.get(), self._v_profit.get()]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch]*4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#f3e8ff')),
                ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#fce7f3')),
                ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#fef3c7')),
                ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#d1fae5')),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Transaction details table
            table_data = [[
                'Date', 'Type', 'Invoice', 'Party', 'Product', 'Batch',
                'Qty', 'P.Rate', 'S.Rate', 'GST', 'P.Cost', 'S.Value', 'Profit'
            ]]
            
            # Collect data from tree
            for item in self.tree.get_children():
                vals = self.tree.item(item, "values")
                tags = self.tree.item(item, "tags")
                
                # Skip total row for now
                if "total" in tags:
                    continue
                
                # Clean values (remove currency symbols and commas)
                cleaned = []
                for v in vals:
                    v_str = str(v).replace("₹", "").replace(",", "").strip()
                    # Truncate long text
                    if len(v_str) > 20:
                        v_str = v_str[:17] + "..."
                    cleaned.append(v_str)
                table_data.append(cleaned)
            
            # Add total row
            for item in self.tree.get_children():
                if "total" in self.tree.item(item, "tags"):
                    vals = self.tree.item(item, "values")
                    cleaned = [str(v).replace("₹", "").replace(",", "").strip() for v in vals]
                    table_data.append(cleaned)
                    break
            
            # Create table with adjusted column widths for landscape
            col_widths = [0.7*inch, 0.7*inch, 0.8*inch, 1*inch, 1.2*inch, 0.6*inch,
                         0.5*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.7*inch, 0.7*inch, 0.7*inch]
            
            details_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style the table
            table_style = TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
                
                # Total row (last row)
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e293b')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 8),
            ])
            
            details_table.setStyle(table_style)
            elements.append(details_table)
            
            # Build PDF
            doc.build(elements)
            
            if messagebox.askyesno("Export Complete",
                                  f"✅ PDF saved to:\n{fp}\n\nOpen file now?",
                                  parent=self):
                import os
                os.startfile(fp)
                
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "ReportLab library is required for PDF export.\n\n"
                "Install it using: pip install reportlab",
                parent=self
            )
        except Exception as e:
            messagebox.showerror("Error", f"PDF export failed: {e}\n\nUse Excel export as alternative.", parent=self)
            import traceback
            traceback.print_exc()

    def kb_refresh(self): self._load()
